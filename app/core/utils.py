#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Utils Module - Các hàm dùng chung cho tất cả các version CapCut
Tách ra để tránh trùng lặp code (Code Duplication)
"""

import json
import os
import csv
import re
import uuid
from pathlib import Path
from copy import deepcopy
from typing import List, Dict, Any

# ==============================================================================
# CÁC HÀM ĐỌC/GHI FILE (FILE I/O)
# ==============================================================================


def load_json_file(file_path: str) -> Dict[str, Any]:
    """Đọc nội dung file JSON."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Lỗi khi đọc file {file_path}: {e}")
        return {}


def save_json_file(data: dict, file_path):
    """Lưu nội dung vào file JSON."""
    try:
        # Hỗ trợ cả Path object và string
        if isinstance(file_path, Path):
            file_path.write_text(
                json.dumps(data, ensure_ascii=False, indent=4), encoding="utf-8"
            )
        else:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"✅ Đã lưu thành công vào file: {file_path}")
    except Exception as e:
        print(f"Lỗi khi lưu file {file_path}: {e}")


# ==============================================================================
# CÁC HÀM XỬ LÝ THỜI GIAN & TEXT
# ==============================================================================


def parse_time_to_microseconds(time_str: str) -> int:
    """Chuyển đổi thời gian từ định dạng MM:SS.mmm sang microseconds"""
    try:
        parts = time_str.split(":")
        minutes = int(parts[0])
        seconds = float(parts[1])
        total_seconds = minutes * 60 + seconds
        return int(total_seconds * 1000000)
    except Exception:
        return 0


def format_timing(milliseconds: int) -> str:
    """Chuyển đổi milliseconds thành định dạng thời gian MM:SS.mmm"""
    total_seconds = milliseconds / 1000
    minutes = int(total_seconds // 60)
    seconds = total_seconds % 60
    return f"{minutes:02d}:{seconds:06.3f}"


def milliseconds_to_srt_time(ms: int) -> str:
    """Chuyển đổi milliseconds sang định dạng thời gian SRT: HH:MM:SS,mmm"""
    total_seconds = ms // 1000
    milliseconds = ms % 1000
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d},{milliseconds:03d}"


def clean_text_from_html(text: str) -> str:
    """Loại bỏ các HTML tags và lấy text thuần túy."""
    if not isinstance(text, str):
        return ""
    # Loại bỏ tags
    text = re.sub(r"<[^>]+>", "", text)
    # Loại bỏ ngoặc vuông thừa
    text = text.replace("[", "").replace("]", "")
    # Xử lý khoảng trắng
    text = " ".join(text.split())
    return text.strip()


def extract_plain_text_from_content(content: str) -> str:
    """Trích xuất text thuần túy từ content (JSON string hoặc text thường)."""
    if not isinstance(content, str):
        return ""
    if content.startswith("{") and content.endswith("}"):
        try:
            content_data = json.loads(content)
            if "text" in content_data:
                return clean_text_from_html(str(content_data["text"]))
        except Exception:
            pass
    return clean_text_from_html(content)


def _tokenize_caption_words(text: str) -> list[str]:
    """Tách câu thành danh sách token đơn giản."""
    if not isinstance(text, str):
        return []
    parts = [p for p in (text or "").split() if p]
    return parts


def is_chinese(text: str) -> bool:
    """Kiểm tra xem text có chứa ký tự tiếng Trung không."""
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def get_cn_texts(data: dict) -> list:
    """Trích xuất danh sách các text có chứa ký tự tiếng Trung từ 'materials > texts'."""
    texts = []
    if "materials" in data and "texts" in data["materials"]:
        for text_item in data["materials"]["texts"]:
            content = text_item.get("content", "")
            if is_chinese(content):
                texts.append(text_item)
    return texts


def extract_text_from_content(content_str: str) -> str:
    """Trích xuất text từ content JSON string"""
    try:
        content_data = json.loads(content_str)
        return content_data.get("text", "")
    except Exception:
        return content_str


# ==============================================================================
# CÁC HÀM TÌM KIẾM TRACKS (FIND TRACKS)
# ==============================================================================


def find_video_tracks(tracks: List[dict]) -> List[dict]:
    """Trả về danh sách các video tracks"""
    return [t for t in tracks if isinstance(t, dict) and t.get("type") == "video"]


def find_video_track(tracks: List[dict]):
    """Tìm track video đầu tiên"""
    video_tracks = find_video_tracks(tracks)
    return video_tracks[0] if video_tracks else None


def find_audio_tracks(tracks: List[dict]) -> List[dict]:
    """Trả về danh sách các audio tracks"""
    return [t for t in tracks if isinstance(t, dict) and t.get("type") == "audio"]


def find_text_tracks(tracks: List[dict]) -> List[dict]:
    """Tìm tất cả các track văn bản (text/caption)."""
    return [t for t in tracks if isinstance(t, dict) and t.get("type") == "text"]


def find_effect_tracks(tracks: List[dict]) -> List[dict]:
    """Tìm tất cả các track hiệu ứng (effect)."""
    return [t for t in tracks if isinstance(t, dict) and t.get("type") == "effect"]


# ==============================================================================
# CÁC HÀM XỬ LÝ LOGIC VIDEO (SPLIT, SYNC)
# ==============================================================================


def get_csv_timing_points(csv_file: str) -> List[int]:
    """Lấy danh sách các điểm thời gian từ CSV."""
    timing_points = []
    try:
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                start_time_str = row.get("start_time", "")
                if start_time_str:
                    start_time_us = parse_time_to_microseconds(start_time_str)
                    if start_time_us >= 0:
                        timing_points.append(start_time_us)
    except Exception as e:
        print(f"Lỗi khi đọc file CSV '{csv_file}': {e}")
        return []
    return sorted(list(set(timing_points)))


def split_segment_by_timing_points(
    segment: Dict[str, Any], timing_points_us: List[int]
) -> List[Dict[str, Any]]:
    """Cắt segment video dựa trên các điểm thời gian."""
    source_timerange = segment.get("source_timerange", {})
    source_start = int(source_timerange.get("start", 0))
    source_duration = int(source_timerange.get("duration", 0))
    source_end = source_start + source_duration

    valid_points = [p for p in timing_points_us if source_start < p < source_end]

    if not valid_points:
        return [deepcopy(segment)]

    all_points = sorted(list(set([source_start] + valid_points + [source_end])))

    new_segments = []
    cursor_us = int(segment.get("target_timerange", {}).get("start", 0))

    for i in range(len(all_points) - 1):
        start_point = all_points[i]
        end_point = all_points[i + 1]

        if start_point >= end_point:
            continue

        duration_us = end_point - start_point

        if duration_us <= 0:
            duration_us = 1

        new_seg = deepcopy(segment)
        new_seg["id"] = str(uuid.uuid4()).upper()
        new_seg["source_timerange"] = {"start": start_point, "duration": duration_us}
        new_seg["target_timerange"] = {"start": cursor_us, "duration": duration_us}

        new_segments.append(new_seg)
        cursor_us += duration_us

    return new_segments


def split_video_track_by_text_timing(video_track: dict, timing_points: List[int]):
    """Chia video track theo timing points."""
    if not video_track or "segments" not in video_track:
        return

    original_segments = video_track.get("segments", [])
    if not original_segments:
        return

    final_segments = []

    for segment in original_segments:
        new_sub_segments = split_segment_by_timing_points(segment, timing_points)
        final_segments.extend(new_sub_segments)

    final_segments.sort(key=lambda s: s.get("target_timerange", {}).get("start", 0))

    for i, segment in enumerate(final_segments):
        segment["track_render_index"] = i
        if "render_timerange" in segment:
            segment["render_timerange"] = {"start": 0, "duration": 0}

    video_track["segments"] = final_segments
    print(f"Đã xử lý và chia video thành {len(final_segments)} segments.")


# ==============================================================================
# CÁC HÀM EXPORT (CSV, SRT, EXCEL)
# ==============================================================================


def save_to_csv(results: List[Dict[str, Any]], filename: str = "translated_texts.csv"):
    """Lưu kết quả ra file CSV"""
    try:
        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = [
                "text",
                "language",
                "id",
                "group_id",
                "start_time",
                "end_time",
                "font_size",
                "text_color",
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for result in results:
                row = {
                    "text": result["text"],
                    "language": result["language"],
                    "id": result["id"],
                    "group_id": result["group_id"],
                    "start_time": "",
                    "end_time": "",
                    "font_size": result["font_size"],
                    "text_color": result["text_color"],
                }

                if result.get("timing_from_subtitle"):
                    timing = result["timing_from_subtitle"]
                    row["start_time"] = format_timing(timing["start_time"])
                    row["end_time"] = format_timing(timing["end_time"])

                writer.writerow(row)

        print(f"Kết quả đã được lưu vào file: {filename}")
    except Exception as e:
        print(f"Lỗi lưu CSV: {e}")


def create_captions_xlsx_if_not_exists(
    results: List[Dict[str, Any]], xlsx_filename: str
):
    """Kiểm tra và tạo file Excel captions."""
    if os.path.isfile(xlsx_filename):
        print(f"ℹ️ File '{xlsx_filename}' đã tồn tại.")
        return

    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠️ Cảnh báo: Thiếu thư viện openpyxl.")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Captions"

        ws["A1"], ws["B1"], ws["C1"], ws["D1"] = (
            "Tiếng Việt",
            "",
            "Phân chia nhân vật",
            "Tiếng Trung",
        )

        for row_idx, item in enumerate(results, start=2):
            ws[f"A{row_idx}"] = item.get("text", "")
            ws[f"B{row_idx}"] = ""
            ws[f"C{row_idx}"] = ""

            chinese_text = ""
            if item.get("timing_from_subtitle") and item["timing_from_subtitle"].get(
                "original_text"
            ):
                chinese_text = item["timing_from_subtitle"]["original_text"]

            ws[f"D{row_idx}"] = chinese_text

        wb.save(xlsx_filename)
        print(f"✅ Đã tạo file Excel: {xlsx_filename}")
    except Exception as e:
        print(f"❌ Lỗi khi tạo Excel: {e}")


def export_to_srt(json_file: str, output_path: str) -> bool:
    """
    Xuất subtitle từ draft_content.json ra file SRT.
    Lấy text và timing từ extra_info > subtitle_fragment_info_list > sentence_list
    (Đây là nguồn dữ liệu chính xác mà CapCut sử dụng để xuất SRT)
    Có fallback logic để lấy từ text tracks nếu không có subtitle fragments.
    """
    try:
        # Đọc file JSON
        json_data = load_json_file(json_file)
        if not json_data:
            print(f"❌ Không thể đọc file '{json_file}'")
            return False

        # Lấy subtitle từ extra_info > subtitle_fragment_info_list
        srt_entries = []

        if (
            "extra_info" in json_data
            and json_data["extra_info"] is not None
            and isinstance(json_data["extra_info"], dict)
            and "subtitle_fragment_info_list" in json_data["extra_info"]
        ):
            subtitle_fragments = json_data["extra_info"]["subtitle_fragment_info_list"]

            for fragment in subtitle_fragments:
                subtitle_cache_info = fragment.get("subtitle_cache_info", "")
                if not subtitle_cache_info:
                    continue

                try:
                    cache_info = json.loads(subtitle_cache_info)
                    if "sentence_list" in cache_info and cache_info["sentence_list"]:
                        for sentence in cache_info["sentence_list"]:
                            # Lấy text từ sentence (text gốc, không phải translation)
                            text = sentence.get("text", "").strip()
                            if not text:
                                continue

                            # Lấy timing từ sentence (đã là milliseconds)
                            start_time_ms = sentence.get("start_time", 0)
                            end_time_ms = sentence.get("end_time", 0)

                            # Đảm bảo timing hợp lệ
                            if start_time_ms >= 0 and end_time_ms > start_time_ms:
                                # Loại bỏ HTML tags nếu có
                                clean_text = clean_text_from_html(text)
                                if clean_text:
                                    srt_entries.append(
                                        {
                                            "start_time": int(start_time_ms),
                                            "end_time": int(end_time_ms),
                                            "text": clean_text,
                                        }
                                    )
                except Exception as e:
                    # Bỏ qua fragment lỗi, tiếp tục với fragment tiếp theo
                    continue

        if not srt_entries:
            print(
                "❌ Không tìm thấy subtitle nào từ extra_info > subtitle_fragment_info_list"
            )
            # Thử fallback: lấy từ text tracks nếu không có subtitle fragments
            print("   Đang thử lấy từ text tracks...")

            # Lấy text từ materials > texts
            materials = json_data.get("materials", {})
            texts_map = {}

            if materials and isinstance(materials, dict) and "texts" in materials:
                for text_item in materials["texts"]:
                    text_id = text_item.get("id", "")
                    if not text_id:
                        continue

                    content = text_item.get("content", "")
                    text_content = extract_plain_text_from_content(content)

                    if not text_content:
                        recognize_text = text_item.get("recognize_text", "")
                        base_content = text_item.get("base_content", "")
                        text_content = clean_text_from_html(
                            recognize_text
                        ) or clean_text_from_html(base_content)

                    if text_content:
                        texts_map[text_id] = text_content.strip()

            # Lấy từ text tracks
            text_segments = []
            tracks = json_data.get("tracks", [])
            for track in tracks:
                if track.get("type") == "text" and "segments" in track:
                    text_segments.extend(track["segments"])

            text_segments.sort(
                key=lambda s: s.get("target_timerange", {}).get("start", 0)
            )

            for segment in text_segments:
                material_id = segment.get("material_id", "")
                target_timerange = segment.get("target_timerange", {})
                start_us = target_timerange.get("start", 0)
                duration_us = target_timerange.get("duration", 0)
                end_us = start_us + duration_us

                text_content = texts_map.get(material_id, "")
                if not text_content:
                    segment_content = segment.get("content", "") or segment.get(
                        "text", ""
                    )
                    if segment_content:
                        text_content = extract_plain_text_from_content(
                            str(segment_content)
                        )

                if text_content:
                    start_ms = start_us // 1000
                    end_ms = end_us // 1000
                    if start_ms >= 0 and end_ms > start_ms:
                        srt_entries.append(
                            {
                                "start_time": start_ms,
                                "end_time": end_ms,
                                "text": text_content,
                            }
                        )

        if not srt_entries:
            print("❌ Không tìm thấy subtitle nào để xuất")
            return False

        # Sắp xếp theo thời gian bắt đầu
        srt_entries.sort(key=lambda x: x["start_time"])

        # Ghi ra file SRT
        try:
            with open(output_path, "w", encoding="utf-8") as f:
                for idx, entry in enumerate(srt_entries, 1):
                    # Số thứ tự
                    f.write(f"{idx}\n")

                    # Thời gian: HH:MM:SS,mmm --> HH:MM:SS,mmm
                    start_time_str = milliseconds_to_srt_time(entry["start_time"])
                    end_time_str = milliseconds_to_srt_time(entry["end_time"])
                    f.write(f"{start_time_str} --> {end_time_str}\n")

                    # Nội dung subtitle
                    f.write(f"{entry['text']}\n")

                    # Dòng trống giữa các entry
                    f.write("\n")

            print(f"✅ Đã xuất {len(srt_entries)} subtitle vào file: {output_path}")
            return True
        except Exception as e:
            print(f"❌ Lỗi khi ghi file SRT: {e}")
            return False

    except Exception as e:
        print(f"❌ Lỗi khi xử lý: {e}")
        import traceback

        traceback.print_exc()
        return False


def export_from_csv(csv_path: str, output_path: str) -> bool:
    """Xuất caption vi-VN từ CSV ra file TXT"""
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            lines = []
            for row in reader:
                if (row.get("language") or "").strip() != "vi-VN":
                    continue
                text_val = (row.get("text") or "").strip()
                if not text_val:
                    continue
                clean = text_val.replace("\r", " ").replace("\n", " ").strip()
                if clean:
                    lines.append(clean)
    except Exception as e:
        print(f"Lỗi khi đọc CSV: {e}")
        return False

    if not lines:
        print("Không tìm thấy caption vi-VN trong CSV để xuất")
        return False

    try:
        with open(output_path, "w", encoding="utf-8") as out:
            for line in lines:
                out.write(line + "\n")
        print(f"Đã xuất {len(lines)} caption vi-VN từ CSV vào: {output_path}")
        return True
    except Exception as e:
        print(f"Lỗi khi ghi file TXT: {e}")
        return False


def export_chinese_from_results(
    results: List[Dict[str, Any]], output_path: str
) -> bool:
    """Xuất caption tiếng Trung từ results ra file TXT"""
    try:
        lines = []
        for item in results:
            # Lấy text tiếng Trung từ timing_from_subtitle
            chinese_text = ""
            if item.get("timing_from_subtitle") and item["timing_from_subtitle"].get(
                "original_text"
            ):
                chinese_text = item["timing_from_subtitle"]["original_text"]

            if chinese_text:
                clean = chinese_text.replace("\r", " ").replace("\n", " ").strip()
                if clean:
                    lines.append(clean)
    except Exception as e:
        print(f"Lỗi khi xử lý results: {e}")
        return False

    if not lines:
        print("Không tìm thấy caption tiếng Trung để xuất")
        return False

    try:
        with open(output_path, "w", encoding="utf-8") as out:
            for line in lines:
                out.write(line + "\n")
        print(f"Đã xuất {len(lines)} caption tiếng Trung vào: {output_path}")
        return True
    except Exception as e:
        print(f"Lỗi khi ghi file TXT: {e}")
        return False


def export_chinese_with_char_count(
    results: List[Dict[str, Any]], output_path: str
) -> bool:
    """Xuất caption tiếng Trung kèm số ký tự mỗi dòng."""
    try:
        lines: List[str] = []
        for item in results:
            chinese_text = ""
            if item.get("timing_from_subtitle") and item["timing_from_subtitle"].get(
                "original_text"
            ):
                chinese_text = item["timing_from_subtitle"]["original_text"]

            if chinese_text:
                clean = chinese_text.replace("\r", " ").replace("\n", " ").strip()
                if clean:
                    char_count = len(clean)
                    # Số ký tự
                    lines.append(f"{clean} ({char_count})")
    except Exception as e:
        print(f"Lỗi khi xử lý results: {e}")
        return False

    if not lines:
        print("Không tìm thấy caption tiếng Trung để xuất")
        return False

    try:
        with open(output_path, "w", encoding="utf-8") as out:
            for line in lines:
                out.write(line + "\n")
        print(
            f"Đã xuất {len(lines)} caption tiếng Trung kèm số ký tự vào: {output_path}"
        )
        return True
    except Exception as e:
        print(f"Lỗi khi ghi file TXT: {e}")
        return False
