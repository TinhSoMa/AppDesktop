#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CapCut Functions Module - Các hàm xử lý CapCut draft
Tách từ mainv7_0.py để tái sử dụng trong CapcutToolv2.py
"""

import json
import uuid
import csv
import re
import os
import logging
from copy import deepcopy
from pathlib import Path
from typing import List, Dict, Any

# Import các hàm dùng chung từ utils
# Import các hàm dùng chung từ utils
try:
    from app.core.utils import (
        load_json_file,
        find_video_track,
        find_audio_tracks,
        find_video_tracks,
        find_text_tracks,
        find_effect_tracks,
        format_timing,
        save_to_csv,
        create_captions_xlsx_if_not_exists,
        split_video_track_by_text_timing,
        get_csv_timing_points,
        export_to_srt,
        milliseconds_to_srt_time,
        _tokenize_caption_words,
        is_chinese,
        export_from_csv,
        export_chinese_from_results,
        export_chinese_with_char_count,
        get_cn_texts,
        extract_text_from_content,
        clean_text_from_html,
    )
except ImportError:
    # Fallback cho trường hợp chạy trực tiếp hoặc cấu trúc khác
    from utils import (
        load_json_file,
        find_video_track,
        find_audio_tracks,
        find_video_tracks,
        find_text_tracks,
        find_effect_tracks,
        format_timing,
        save_to_csv,
        create_captions_xlsx_if_not_exists,
        split_video_track_by_text_timing,
        get_csv_timing_points,
        export_to_srt,
        milliseconds_to_srt_time,
        _tokenize_caption_words,
        is_chinese,
        export_from_csv,
        export_chinese_from_results,
        export_chinese_with_char_count,
        get_cn_texts,
        extract_text_from_content,
    )


# --- Hàm phụ trợ cho xử lý TEXT (riêng cho version này) ---
# get_cn_texts và extract_text_from_content đã được chuyển sang utils.py

# ==============================================================================
# CÁC HÀM CHỨC NĂNG CHÍNH
# ==============================================================================


def get_translated_texts_with_timing(json_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Lấy thông tin text đã được translate từ materials > texts
    Bao gồm text tiếng Việt và timing
    CẢI TIẾN: Xử lý cả content JSON và string thường
    """
    results = []

    # Lấy subtitle timing từ extra_info
    subtitle_timings = []
    if (
        "extra_info" in json_data
        and "subtitle_fragment_info_list" in json_data["extra_info"]
    ):
        for fragment in json_data["extra_info"]["subtitle_fragment_info_list"]:
            if "subtitle_cache_info" in fragment and fragment["subtitle_cache_info"]:
                try:
                    cache_info = json.loads(fragment["subtitle_cache_info"])
                    if "sentence_list" in cache_info:
                        for sentence in cache_info["sentence_list"]:
                            if (
                                "bilingual_lan" in sentence
                                and sentence["bilingual_lan"] == "vi-VN"
                            ):
                                start_time = sentence.get("start_time", 0)
                                end_time = sentence.get("end_time", 0)
                                original_text = sentence.get("text", "")
                                translation_text = sentence.get("translation_text", "")

                                subtitle_timings.append(
                                    {
                                        "start_time": start_time,
                                        "end_time": end_time,
                                        "original_text": original_text,
                                        "translation_text": translation_text,
                                    }
                                )
                except:
                    continue

    # Lấy text từ materials > texts
    if "materials" in json_data and "texts" in json_data["materials"]:
        for text_item in json_data["materials"]["texts"]:
            # Chỉ lấy text có language là vi-VN
            if text_item.get("language") == "vi-VN":
                # Lấy text từ content - XỬ LÝ CẢ 2 DẠNG
                content = text_item.get("content", "")
                if content.startswith("{") and content.endswith("}"):
                    try:
                        content_data = json.loads(content)
                        text = content_data.get("text", "")
                    except:
                        text = content
                else:
                    text = content

                # Lấy text từ recognize_text nếu có
                recognize_text = text_item.get("recognize_text", "")

                # Lấy text từ base_content nếu có
                base_content = text_item.get("base_content", "")

                # Ưu tiên text từ content, sau đó recognize_text, cuối cùng base_content
                final_text = text or recognize_text or base_content

                if final_text:
                    # Tìm timing tương ứng từ subtitle
                    timing_info = None
                    # Lấy timing theo thứ tự xuất hiện
                    if len(results) < len(subtitle_timings):
                        timing_info = subtitle_timings[len(results)]

                    # Lấy timing từ current_words nếu có
                    current_words = text_item.get("current_words", {})
                    start_times = current_words.get("start_time", [])
                    end_times = current_words.get("end_time", [])

                    result = {
                        "text": final_text,
                        "language": text_item.get("language", ""),
                        "id": text_item.get("id", ""),
                        "group_id": text_item.get("group_id", ""),
                        "timing_from_subtitle": timing_info,
                        "timing_from_current_words": (
                            {"start_times": start_times, "end_times": end_times}
                            if start_times or end_times
                            else None
                        ),
                        "font_size": text_item.get("font_size", 0),
                        "text_color": text_item.get("text_color", ""),
                        "alignment": text_item.get("alignment", 0),
                    }

                    results.append(result)

    return results


# format_timing đã được chuyển sang utils.py


def print_results(results: List[Dict[str, Any]]):
    """In kết quả ra màn hình"""
    print(f"Tìm thấy {len(results)} text đã được translate:\n")
    print("=" * 80)

    for i, result in enumerate(results, 1):
        print(f"\n{i}. Text ID: {result['id']}")
        print(f"   Group ID: {result['group_id']}")
        print(f"   Text: {result['text']}")
        print(f"   Language: {result['language']}")
        print(f"   Font Size: {result['font_size']}")
        print(f"   Text Color: {result['text_color']}")
        print(f"   Alignment: {result['alignment']}")

        # Timing từ subtitle
        if result["timing_from_subtitle"]:
            timing = result["timing_from_subtitle"]
            start_time = format_timing(timing["start_time"])
            end_time = format_timing(timing["end_time"])
            print(f"   Timing (từ subtitle): {start_time} - {end_time}")
            print(f"   Original text: {timing['original_text']}")
            if timing["translation_text"]:
                print(f"   Translation text: {timing['translation_text']}")

        # Timing từ current_words
        if result["timing_from_current_words"]:
            timing = result["timing_from_current_words"]
            if timing["start_times"] and timing["end_times"]:
                print(f"   Timing (từ current_words):")
                for j, (start, end) in enumerate(
                    zip(timing["start_times"], timing["end_times"])
                ):
                    start_time = format_timing(start)
                    end_time = format_timing(end)
                    print(f"     Word {j+1}: {start_time} - {end_time}")

        print("-" * 80)


# save_to_csv đã được chuyển sang utils.py

# ==============================================================================
# CẤU HÌNH STYLE ĐƯỢC NHÚNG TRỰC TIẾP TỪ DRAFT_CONTENT_FIX.JSON
# ==============================================================================


def get_hardcoded_style_config() -> Dict[str, Any]:
    """
    Trả về cấu hình style đã được sao chép chính xác từ file mẫu thành công.
    CẬP NHẬT: Sử dụng cấu trúc thực tế từ saukhisuastyle.json
    """
    return {
        "text_style_template": {
            "add_type": 0,
            "alignment": 1,
            "background_alpha": 1.0,
            "background_color": "#000000",
            "background_fill": "",
            "background_height": 0.14,
            "background_horizontal_offset": 0.0,
            "background_round_radius": 0.0,
            "background_style": 0,
            "background_vertical_offset": 0.0,
            "background_width": 0.14,
            "base_content": "",
            "bold_width": 0.0,
            "border_alpha": 1.0,
            "border_color": "#ffffff",
            "border_width": 0.08,
            "caption_template_info": {
                "category_id": "",
                "category_name": "",
                "effect_id": "",
                "is_new": False,
                "path": "",
                "request_id": "",
                "resource_id": "",
                "resource_name": "",
                "source_platform": 0,
                "third_resource_id": "",
            },
            "check_flag": 47,
            "combo_info": {"text_templates": []},
            "cutoff_postfix": "",
            "enable_path_typesetting": False,
            "fixed_height": -1.0,
            "fixed_width": -1.0,
            "font_category_id": "",
            "font_category_name": "",
            "font_id": "",
            "font_name": "",
            "font_path": "C:/Users/congt/AppData/Local/CapCut/User Data/Cache/effect/7535354391860120848/864dc8c9046e5845640daba0bdeab144/font.ttf",
            "font_resource_id": "7535354391860120848",
            "font_size": 8.0,  # ✅ ĐÃ XÁC NHẬN: 8.0
            "font_source_platform": 1,
            "font_team_id": "",
            "font_third_resource_id": "",
            "font_title": "none",  # ✅ ĐÃ XÁC NHẬN: "none" từ file mẫu
            "font_url": "",
            "force_apply_line_max_width": False,
            "global_alpha": 1.0,
            "has_shadow": False,
            "initial_scale": 1.0,
            "inner_padding": -1.0,
            "is_lyric_effect": False,
            "is_rich_text": False,
            "is_words_linear": False,
            "italic_degree": 0,
            "ktv_color": "",
            "language": "vi-VN",
            "layer_weight": 1,
            "letter_spacing": 0.0,
            "line_feed": 1,
            "line_max_width": 0.82,
            "line_spacing": 0.02,
            "lyric_group_id": "",
            "lyrics_template": {
                "category_id": "",
                "category_name": "",
                "effect_id": "",
                "panel": "",
                "path": "",
                "request_id": "",
                "resource_id": "",
                "resource_name": "",
            },
            "multi_language_current": "none",
            "name": "",
            "offset_on_path": 0.0,
            "oneline_cutoff": False,
            "operation_type": 0,
            "original_size": [],
            "preset_category": "",
            "preset_category_id": "",
            "preset_has_set_alignment": False,
            "preset_id": "",
            "preset_index": 0,
            "preset_name": "",
            "recognize_task_id": "",
            "recognize_type": 0,
            "relevance_segment": [],
            "shadow_alpha": 0.9,
            "shadow_angle": -45.0,
            "shadow_color": "#000000",
            "shadow_distance": 5.0,
            "shadow_point": {"x": 0.6363961030678928, "y": -0.6363961030678928},
            "shadow_smoothing": 0.45,
            "shape_clip_x": False,
            "shape_clip_y": False,
            "source_from": "",
            "ssml_content": "",
            "style_name": "",
            "sub_template_id": -1,
            "sub_type": 5,
            "subtitle_keywords": None,
            "subtitle_keywords_config": None,
            "subtitle_template_original_fontsize": 0.0,
            "text_alpha": 1.0,
            "text_color": "#ffffff",
            "text_curve": None,
            "text_exceeds_path_process_type": 0,
            "text_loop_on_path": False,
            "text_preset_resource_id": "",
            "text_size": 30,
            "text_to_audio_ids": [],
            "text_typesetting_path_index": 0,
            "text_typesetting_paths": None,
            "text_typesetting_paths_file": "",
            "translate_original_text": "",
            "tts_auto_update": False,
            "type": "subtitle",  # QUAN TRỌNG: Sử dụng "subtitle" thay vì "text"
            "typesetting": 0,
            "underline": False,
            "underline_offset": 0.22,
            "underline_width": 0.05,
            "use_effect_default_color": True,
        },
        "font_template": {
            "category_id": "favoured",
            "category_name": "Yêu thích",
            "effect_id": "7535354391860120848",
            "file_uri": "",
            "id": "470A81BA-465E-4a20-992C-C4FEFA0E296C",  # ID từ file mẫu
            "path": "C:/Users/congt/AppData/Local/CapCut/User Data/Cache/effect/7535354391860120848/864dc8c9046e5845640daba0bdeab144/font.ttf",
            "request_id": "",
            "resource_id": "7535354391860120848",
            "source_platform": 1,
            "team_id": "",
            "third_resource_id": "",
            "title": "Tiếng Việt",  # ✅ ĐÃ XÁC NHẬN: "Tiếng Việt"
        },
        "content_styles_template": {
            "styles": [
                {
                    "fill": {
                        "alpha": 1.0,
                        "content": {
                            "render_type": "solid",
                            "solid": {"alpha": 1.0, "color": [1.0, 1.0, 1.0]},
                        },
                    },
                    "font": {
                        "id": "7535354391860120848",
                        "path": "C:/Users/congt/AppData/Local/CapCut/User Data/Cache/effect/7535354391860120848/864dc8c9046e5845640daba0bdeab144/font.ttf",
                    },
                    "range": [0, 0],  # Sẽ được cập nhật theo độ dài text
                    "shadows": [
                        {
                            "alpha": 1.0,
                            "angle": 0.0,
                            "content": {
                                "render_type": "solid",
                                "solid": {"alpha": 1.0, "color": [0.0, 0.0, 0.0]},
                            },
                            "diffuse": 0.0833333358168602,
                            "distance": 0.0,
                            "feather": 0.15,
                        }
                    ],
                    "size": 8,
                    "useLetterColor": True,
                }
            ],
            "text": "",  # Sẽ được cập nhật với nội dung thực tế
        },
    }


def apply_style_to_texts(data_to_update: Dict) -> bool:
    """
    Áp dụng style đã được nhúng vào TẤT CẢ text vi-VN trong draft.
    CẬP NHẬT: Áp dụng cho toàn bộ text của tất cả các track text, không chỉ riêng vi-VN.
    """
    style_config = get_hardcoded_style_config()
    text_template = style_config["text_style_template"]
    font_template = style_config["font_template"]
    content_styles_template = style_config["content_styles_template"]

    # Cập nhật fonts trong materials
    materials = data_to_update.setdefault("materials", {})
    fonts = materials.setdefault("fonts", [])

    if not fonts:
        fonts.append(font_template)
        logging.info(f"✅ Đã thêm font '{font_template['title']}' vào materials.")
    else:
        # Kiểm tra xem font đã tồn tại chưa
        font_exists = any(f.get("id") == font_template["id"] for f in fonts)
        if not font_exists:
            fonts.append(font_template)
            logging.info(f"✅ Đã thêm font '{font_template['title']}' vào materials.")

    # Lấy danh sách TẤT CẢ text (không chỉ vi-VN)
    texts = materials.get("texts", [])

    if not texts:
        logging.warning("⚠️ Không tìm thấy text nào để áp dụng style.")
        return False

    # Tạo một danh sách mới để chứa các text đã được nâng cấp
    new_styled_texts = []

    for text_item in texts:
        # Áp dụng style cho TẤT CẢ text
        # 1. Giữ lại các thông tin gốc quan trọng
        original_id = text_item.get("id")
        original_group_id = text_item.get("group_id")
        original_language = text_item.get("language", "")

        # Lấy nội dung text từ cả 2 dạng (string thường hoặc JSON string)
        content_str = text_item.get("content", "")
        if content_str.startswith("{") and content_str.endswith("}"):
            try:
                content_data = json.loads(content_str)
                original_text_content = content_data.get("text", "")
            except:
                original_text_content = content_str
        else:
            original_text_content = content_str

        # 2. Tạo một đối tượng text mới từ template
        new_text_obj = deepcopy(text_template)

        # 3. Tạo content mới với cấu trúc styles phức tạp
        new_content_styles = deepcopy(content_styles_template)
        new_content_styles["text"] = original_text_content
        # Cập nhật range theo độ dài text
        text_length = len(original_text_content)
        new_content_styles["styles"][0]["range"] = [0, text_length]

        # 4. Điền lại thông tin gốc vào đối tượng mới
        new_text_obj["id"] = original_id
        new_text_obj["group_id"] = original_group_id
        new_text_obj["language"] = original_language  # Giữ nguyên ngôn ngữ gốc
        new_text_obj["content"] = json.dumps(new_content_styles, ensure_ascii=False)
        new_text_obj["recognize_text"] = original_text_content
        new_text_obj.setdefault("words", {})["text"] = _tokenize_caption_words(
            original_text_content
        )
        new_text_obj["fonts"] = [deepcopy(font_template)]

        # 5. Thêm đối tượng đã nâng cấp vào danh sách mới
        new_styled_texts.append(new_text_obj)

    # 6. Thay thế toàn bộ danh sách text cũ bằng danh sách mới đã nâng cấp
    materials["texts"] = new_styled_texts

    logging.info(f"✅ Đã áp dụng style thành công cho {len(texts)} text (tất cả ngôn ngữ).")
    return True


# create_captions_xlsx_if_not_exists đã được chuyển sang utils.py


# export_chinese_from_results, export_chinese_with_char_count, export_to_srt, export_from_csv đã được chuyển sang utils.py
# Các hàm này đã được import từ utils.py ở đầu file


def replace_vi_texts_in_draft_from_xlsx(
    draft_path: str,
    xlsx_path: str,
    sheet: str | None = None,
    column: str = "B",
    start_row: int = 2,
    offset: int = 0,
    map_by_index: bool = True,
) -> bool:
    """
    Đọc các caption từ file Excel (mặc định cột B) và thay thế tuần tự vào
    materials > texts (language = vi-VN) trong draft_content.json.
    """
    if not os.path.isfile(draft_path):
        logging.error(f"Không tìm thấy file draft: {draft_path}")
        return False
    if not os.path.isfile(xlsx_path):
        logging.error(f"Không tìm thấy file Excel: {xlsx_path}")
        return False

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        logging.error("Thiếu thư viện openpyxl. Hãy cài đặt: pip install openpyxl")
        return False

    # Đọc JSON draft
    try:
        with open(draft_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logging.error(f"Lỗi khi đọc JSON: {e}")
        return False

    # Đọc Excel
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    except Exception as e:
        logging.error(f"Lỗi khi đọc Excel: {e}")
        return False

    # Thu thập dòng từ cột chỉ định
    col_letter = (column or "B").strip() or "B"
    try:
        from openpyxl.utils import column_index_from_string  # type: ignore

        col_idx = int(column_index_from_string(col_letter))
    except Exception:
        logging.error(f"Cột không hợp lệ: {column}")
        return False

    if start_row < 1:
        start_row = 1

    new_lines: list[str] = []
    try:
        max_row = ws.max_row or 0
        for r in range(start_row, max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            text_val = "" if val is None else str(val)
            clean = text_val.replace("\r", " ").replace("\n", " ").strip()
            if clean:
                new_lines.append(clean)
    except Exception as e:
        logging.error(f"Lỗi khi duyệt cột Excel: {e}")
        return False

    # Thực hiện thay thế theo chỉ số (materials > texts > 0..n) như yêu cầu
    materials = data.get("materials") or {}
    texts = materials.get("texts") or []
    if not isinstance(texts, list) or not texts:
        logging.warning("Không tìm thấy materials > texts trong draft hoặc danh sách rỗng")
        return False

    if offset < 0:
        offset = 0
    if map_by_index:
        start_index = offset
        pair_count = min(len(texts) - start_index, len(new_lines))
        index_resolver = lambda i: start_index + i
    else:
        vi_indexes = []
        for idx, t in enumerate(texts):
            if isinstance(t, dict) and (t.get("language") or "").strip() == "vi-VN":
                vi_indexes.append(idx)
        if not vi_indexes:
            logging.warning("Không tìm thấy texts vi-VN để thay thế")
            return False
        vi_slice = vi_indexes[offset:]
        pair_count = min(len(vi_slice), len(new_lines))
        index_resolver = lambda i: vi_slice[i]
    if pair_count == 0:
        logging.warning("Excel không có dữ liệu để thay thế")
        return False


    updates = 0
    for i in range(pair_count):
        idx = index_resolver(i)
        t = texts[idx] if 0 <= idx < len(texts) else None
        if not isinstance(t, dict):
            continue
        new_text = new_lines[i]

        content_val = t.get("content")
        updated_content = None
        if isinstance(content_val, str) and content_val:
            try:
                parsed = json.loads(content_val)
                if isinstance(parsed, dict):
                    parsed["text"] = new_text
                    updated_content = json.dumps(parsed, ensure_ascii=False)
            except Exception:
                updated_content = new_text
        else:
            updated_content = new_text

        t["content"] = updated_content
        t["recognize_text"] = new_text

        # Cập nhật words theo yêu cầu
        words_obj = t.get("words")
        if not isinstance(words_obj, dict):
            words_obj = {}
            t["words"] = words_obj
        words_obj["text"] = _tokenize_caption_words(new_text)
        words_obj["start_time"] = []
        words_obj["end_time"] = []

        updates += 1

    data.setdefault("materials", {})["texts"] = texts

    # Backup
    try:
        if os.path.isfile(draft_path):
            bak_path = draft_path + ".bak"
            with open(draft_path, "r", encoding="utf-8") as f_in, open(
                bak_path, "w", encoding="utf-8"
            ) as f_out:
                f_out.write(f_in.read())
            logging.info(f"Đã tạo backup: {bak_path}")
    except Exception as e:
        logging.warning(f"Cảnh báo: không thể tạo backup: {e}")

    logging.info(
        f"Đã chuẩn bị cập nhật {updates} texts trong draft từ Excel (offset={offset}, by_index={map_by_index})"
    )
    if len(new_lines) > pair_count:
        logging.info(
            f"Lưu ý: còn {len(new_lines) - pair_count} dòng trong Excel chưa dùng (vượt số lượng texts vi-VN)"
        )

    # Lưu file với nội dung đã cập nhật (KHÔNG áp dụng style)
    try:
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        logging.info(f"✅ Đã lưu file vào: {draft_path}")
        return True
    except Exception as e:
        logging.error(f"❌ Lỗi khi lưu file {draft_path}: {e}")
        return False


def get_srt_timing_points(srt_file: str) -> List[int]:
    """
    Đọc file SRT và trả về danh sách thời gian bắt đầu (ms) của từng subtitle.
    """
    timing_points = []
    try:
        with open(srt_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Tách các entry SRT (mỗi entry bắt đầu bằng số, kết thúc bằng dòng trống)
        entries = content.strip().split('\n\n')
        
        for entry in entries:
            lines = entry.strip().split('\n')
            if len(lines) >= 2:
                # Dòng thứ 2 là timing: "00:00:01,000 --> 00:00:05,000"
                timing_line = lines[1]
                if '-->' in timing_line:
                    start_str = timing_line.split('-->')[0].strip()
                    # Chuyển "00:00:01,000" thành ms
                    start_ms = time_to_ms_srt(start_str)
                    timing_points.append(start_ms)
    except Exception as e:
        print(f"Lỗi khi đọc file SRT: {e}")
    
    return timing_points


def time_to_ms_srt(time_str: str) -> int:
    """
    Chuyển thời gian SRT "HH:MM:SS,mmm" thành mili giây.
    """
    try:
        h, m, s_ms = time_str.split(':')
        s, ms = s_ms.split(',')
        return int(h) * 3600000 + int(m) * 60000 + int(s) * 1000 + int(ms)
    except:
        return 0


def split_video_by_srt_timing(
    json_data: Dict[str, Any], srt_file: str
):
    """CHỨC NĂNG: Chia video dựa trên thời gian từ file SRT"""
    timing_points = get_srt_timing_points(srt_file)
    if not timing_points:
        logging.warning("Không có thông tin timing từ file SRT để chia video.")
        return False

    tracks = json_data.get("tracks", [])
    video_track = find_video_track(tracks)
    if not video_track:
        logging.warning("Không tìm thấy video track.")
        return False

    logging.info(f"Số điểm chia video từ SRT: {len(timing_points)}")

    backup_path = Path("draft_content.json.bak")
    if Path("draft_content.json").exists():
        backup_path.write_text(
            Path("draft_content.json").read_text(encoding="utf-8"), encoding="utf-8"
        )
        logging.info(f"Đã tạo backup: {backup_path}")

    split_video_track_by_text_timing(video_track, timing_points)

    with open("draft_content.json", "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)

    logging.info("Đã cập nhật draft_content.json với video đã được chia.")
    return True


def split_texts_into_multiple_tracks_by_character_index(
    json_data: Dict[str, Any], xlsx_file: str = "captions.xlsx"
) -> bool:
    """
    Phân chia các text segments hiện có thành nhiều track mới
    dựa trên chỉ số nhân vật trong cột 'Phân chia nhân vật' từ file Excel.
    """
    logging.info(f"\n--- CHỨC NĂNG: CHIA TEXT THÀNH NHIỀU TRACK THEO NHÂN VẬT ---")

    if not os.path.isfile(xlsx_file):
        logging.error(f"❌ Không tìm thấy file Excel: {xlsx_file}")
        return False

    try:
        from openpyxl import load_workbook
    except ImportError:
        logging.error("❌ Thiếu thư viện openpyxl. Hãy cài đặt: pip install openpyxl")
        return False

    # 1. Đọc dữ liệu từ Excel
    try:
        wb = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        logging.error(f"❌ Lỗi khi đọc Excel: {e}")
        return False

    # Thu thập dữ liệu từ cột C (Phân chia nhân vật)
    character_indices = []
    try:
        max_row = ws.max_row or 0
        for r in range(2, max_row + 1):  # Bắt đầu từ dòng 2
            cell = ws.cell(row=r, column=3)  # Cột C
            val = cell.value
            char_index_str = "" if val is None else str(val).strip()

            try:
                char_index = int(char_index_str) if char_index_str else 0
            except ValueError:
                char_index = 0  # Không phải số, xếp vào nhóm 0

            character_indices.append(char_index)
    except Exception as e:
        logging.error(f"❌ Lỗi khi đọc cột phân chia nhân vật: {e}")
        return False

    # 2. Thu thập các text segment hiện có từ draft
    tracks = json_data.get("tracks", [])
    existing_text_tracks = find_text_tracks(tracks)

    if not existing_text_tracks or not existing_text_tracks[0].get("segments"):
        logging.warning("❌ Không tìm thấy track text nào trong draft để phân chia.")
        return False

    original_segments = existing_text_tracks[0]["segments"]

    if len(original_segments) != len(character_indices):
        logging.warning(
            f"⚠️ Cảnh báo: Số lượng text segment ({len(original_segments)}) và số dòng Excel ({len(character_indices)}) không khớp."
        )
        logging.warning(
            f"   Chỉ xử lý tối đa {min(len(original_segments), len(character_indices))} segments."
        )
        max_map = min(len(original_segments), len(character_indices))
    else:
        max_map = len(original_segments)

    # 3. Phân loại segments theo chỉ số nhân vật
    character_tracks: Dict[int, List[Dict[str, Any]]] = {}

    for i in range(max_map):
        char_index = character_indices[i]
        segment = original_segments[i]

        # Nhóm các segment theo chỉ số nhân vật
        character_tracks.setdefault(char_index, []).append(segment)

    if not character_tracks:
        logging.warning("⚠️ Không có text nào để phân chia.")
        return False

    # 4. Xây dựng lại danh sách tracks
    new_tracks = [t for t in tracks if t.get("type") != "text"]

    # Tạo track mới cho mỗi nhân vật
    all_char_indices = sorted(character_tracks.keys())

    for char_index in all_char_indices:
        segments = character_tracks[char_index]

        # Sắp xếp lại segments theo thời gian bắt đầu
        segments.sort(key=lambda s: s.get("target_timerange", {}).get("start", 0))

        # Tạo tên track
        if char_index == 0:
            track_name = "Text-Other"
        else:
            track_name = f"Text-Char-{char_index}"

        # Đánh lại track_render_index
        for i, seg in enumerate(segments):
            seg["track_render_index"] = i

        # Tạo track mới
        new_text_track = deepcopy(existing_text_tracks[0])
        new_text_track["id"] = str(uuid.uuid4()).upper()
        new_text_track["segments"] = segments
        new_text_track["flag"] = 0

        new_tracks.append(new_text_track)
        logging.info(f"✅ Đã tạo track '{track_name}' với {len(segments)} segments.")

    # 5. Cập nhật draft
    json_data["tracks"] = new_tracks

    # Tạo backup trước khi lưu
    backup_path = "draft_content.json.bak"
    try:
        if os.path.exists("draft_content.json"):
            if os.path.exists(backup_path):
                os.remove(backup_path)
            os.rename("draft_content.json", backup_path)
            logging.info(f"💾 Đã tạo backup: {backup_path}")
    except Exception as e:
        logging.warning(f"⚠️ Cảnh báo: Không thể tạo backup: {e}")

    # Lưu file
    try:
        with open("draft_content.json", "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)
        logging.info(
            f"✅ Đã lưu file 'draft_content.json' với các track text đã được phân chia."
        )
        return True
    except Exception as e:
        logging.error(f"❌ Lỗi khi lưu file: {e}")
        return False


def split_video_and_slow_down_for_audio(
    json_data: Dict[str, Any], video_start_index: int = 0
) -> bool:
    """CHỨC NĂNG 2: Đồng bộ, làm chậm video theo audio và tái xây dựng timeline.

    LOGIC TEXT TIMING:
    - Thời gian bắt đầu text = thời gian bắt đầu audio hiện tại
    - Thời gian kết thúc text = thời gian bắt đầu audio tiếp theo
    - Text cuối cùng sẽ có thời gian kết thúc = thời gian kết thúc audio cuối cùng
    """
    tracks = json_data.get("tracks", [])
    if not tracks:
        logging.error("❌ Không tìm thấy tracks trong file.")
        return False

    if "materials" not in json_data:
        json_data["materials"] = {}
    if "speeds" not in json_data["materials"]:
        json_data["materials"]["speeds"] = []

    materials = json_data["materials"]
    all_speeds_map = {s["id"]: s for s in materials.get("speeds", [])}

    def get_start_time(seg):
        return int(seg.get("target_timerange", {}).get("start", float("inf")))

    all_video_segments = sorted(
        [
            seg
            for track in find_video_tracks(tracks)
            for seg in track.get("segments", [])
        ],
        key=get_start_time,
    )
    all_audio_segments = sorted(
        [
            seg
            for track in find_audio_tracks(tracks)
            for seg in track.get("segments", [])
        ],
        key=get_start_time,
    )
    all_text_segments = sorted(
        [
            seg
            for track in find_text_tracks(tracks)
            for seg in track.get("segments", [])
        ],
        key=get_start_time,
    )

    all_effect_tracks = find_effect_tracks(tracks)
    effect_material_id = None
    if all_effect_tracks and all_effect_tracks[0].get("segments"):
        effect_material_id = all_effect_tracks[0]["segments"][0].get("material_id")

    if not all_video_segments:
        logging.error("❌ Không tìm thấy video segment nào.")
        return False
    if video_start_index >= len(all_video_segments):
        logging.error(
            f"❌ Lỗi: Bạn muốn bắt đầu từ video số {video_start_index + 1}, nhưng chỉ có {len(all_video_segments)} video."
        )
        return False

    final_video_segs, final_audio_segs, final_text_segs, final_effect_segs = (
        [],
        [],
        [],
        [],
    )
    cursor_us = get_start_time(all_video_segments[0]) if all_video_segments else 0
    logging.info(f"Timeline sẽ được xây dựng lại bắt đầu từ mốc: {cursor_us / 1000:.0f}ms")

    sync_count = min(
        len(all_video_segments) - video_start_index, len(all_audio_segments)
    )
    video_cursor = 0

    for i in range(video_start_index):
        video_seg = deepcopy(all_video_segments[i])
        video_seg["target_timerange"]["start"] = cursor_us
        duration = int(video_seg.get("target_timerange", {}).get("duration", 0))
        cursor_us += duration
        final_video_segs.append(video_seg)
        video_cursor += 1

    logging.info(f"Tìm thấy {sync_count} cặp video-audio để đồng bộ...")
    for i in range(sync_count):
        video_seg = deepcopy(all_video_segments[video_cursor])
        audio_seg = deepcopy(all_audio_segments[i])
        text_seg = (
            deepcopy(all_text_segments[i]) if i < len(all_text_segments) else None
        )
        source_duration_us = int(
            video_seg.get("source_timerange", {}).get("duration", 0)
        )
        if source_duration_us == 0:
            source_duration_us = int(
                video_seg.get("target_timerange", {}).get("duration", 0)
            )

        audio_duration_us = int(
            audio_seg.get("target_timerange", {}).get("duration", 0)
        )
        final_video_duration_us = source_duration_us
        new_speed = 1.0

        # SỬA LỖI: Chỉ làm chậm video khi audio dài hơn video
        # Không làm nhanh video khi audio ngắn hơn
        if audio_duration_us > source_duration_us and source_duration_us > 0:
            new_speed = float(source_duration_us) / float(audio_duration_us)
            new_speed = max(0.1, new_speed)
            final_video_duration_us = audio_duration_us
            logging.info(
                f"🔹 Video {video_cursor + 1} được làm chậm thành {audio_duration_us/1000000:.2f}s (tốc độ: {new_speed:.3f}x)"
            )
        else:
            # Giữ nguyên tốc độ và thời lượng video khi audio không dài hơn
            logging.info(
                f"🔹 Video {video_cursor + 1} giữ nguyên tốc độ (audio: {audio_duration_us/1000000:.2f}s, video: {source_duration_us/1000000:.2f}s)"
            )

        video_seg["speed"] = new_speed
        video_seg["target_timerange"]["start"] = cursor_us
        video_seg["target_timerange"]["duration"] = final_video_duration_us
        if "render_timerange" in video_seg:
            video_seg["render_timerange"] = {"start": 0, "duration": 0}

        # THAY ĐỔI LOGIC CUỐI CÙNG: Luôn tạo mới speed material
        existing_speed_material_id = next(
            (
                ref_id
                for ref_id in video_seg.get("extra_material_refs", [])
                if ref_id in all_speeds_map
            ),
            None,
        )

        # Gỡ bỏ liên kết cũ (nếu có) để làm sạch
        if existing_speed_material_id and "extra_material_refs" in video_seg:
            video_seg["extra_material_refs"].remove(existing_speed_material_id)

        # Nếu tốc độ không phải 1.0, LUÔN TẠO MỚI speed material và liên kết lại
        if new_speed != 1.0:
            new_speed_material = {
                "id": str(uuid.uuid4()).upper(),
                "speed": new_speed,
                "type": "speed",
                "curve_speed": None,
                "mode": 0,
            }
            materials["speeds"].append(new_speed_material)
            all_speeds_map[new_speed_material["id"]] = new_speed_material
            if "extra_material_refs" not in video_seg:
                video_seg["extra_material_refs"] = []
            video_seg["extra_material_refs"].append(new_speed_material["id"])

        final_video_segs.append(video_seg)
        audio_seg["target_timerange"]["start"] = cursor_us
        audio_seg["target_timerange"]["duration"] = audio_duration_us
        final_audio_segs.append(audio_seg)

        # XỬ LÝ TEXT: Toàn bộ text bám audio (bắt đầu = audio start, kết thúc = audio end)
        if text_seg:
            text_seg["target_timerange"]["start"] = cursor_us
            text_seg["target_timerange"]["duration"] = audio_duration_us
            final_text_segs.append(text_seg)
            logging.info(
                f"🔹 Text {i + 1}: {cursor_us/1000000:.2f}s - {(cursor_us + audio_duration_us)/1000000:.2f}s (duration: {audio_duration_us/1000000:.2f}s)"
            )

        # XỬ LÝ EFFECT: Đồng bộ theo audio, nếu có audio tiếp theo còn dư thì cộng thêm 0.2s
        if effect_material_id:
            effect_duration_us = audio_duration_us

            # Kiểm tra xem có audio tiếp theo và còn dư không
            if i + 1 < len(all_audio_segments):
                next_audio_seg = all_audio_segments[i + 1]
                next_audio_duration = int(
                    next_audio_seg.get("target_timerange", {}).get("duration", 0)
                )

                # Nếu audio tiếp theo có duration > 0 (còn dư), cộng thêm 0.2s vào effect
                if next_audio_duration > 0:
                    extra_duration_us = 200000  # 0.2 giây = 200000 microseconds
                    effect_duration_us += extra_duration_us
                    logging.info(
                        f"🔹 Effect {i + 1}: Đã cộng thêm 0.2s (duration: {effect_duration_us/1000000:.2f}s)"
                    )

            new_effect_seg = {
                "id": str(uuid.uuid4()).upper(),
                "material_id": effect_material_id,
                "target_timerange": {
                    "start": cursor_us,
                    "duration": effect_duration_us,
                },
            }
            final_effect_segs.append(new_effect_seg)

        cursor_us += final_video_duration_us
        video_cursor += 1

    for i in range(video_cursor, len(all_video_segments)):
        video_seg = deepcopy(all_video_segments[i])
        video_seg["target_timerange"]["start"] = cursor_us
        duration = int(video_seg.get("target_timerange", {}).get("duration", 0))
        cursor_us += duration
        final_video_segs.append(video_seg)

    new_tracks = [
        t for t in tracks if t.get("type") not in ["video", "audio", "text", "effect"]
    ]
    if final_video_segs:
        new_tracks.insert(
            0,
            {
                "type": "video",
                "segments": final_video_segs,
                "id": str(uuid.uuid4()).upper(),
                "attribute": 0,
                "flag": 0,
            },
        )
    if final_audio_segs:
        new_tracks.append(
            {
                "type": "audio",
                "segments": final_audio_segs,
                "id": str(uuid.uuid4()).upper(),
                "attribute": 0,
                "flag": 0,
            }
        )
    if final_text_segs:
        new_tracks.append(
            {
                "type": "text",
                "segments": final_text_segs,
                "id": str(uuid.uuid4()).upper(),
                "attribute": 0,
                "flag": 0,
            }
        )
    if final_effect_segs:
        new_tracks.append(
            {
                "type": "effect",
                "segments": final_effect_segs,
                "id": str(uuid.uuid4()).upper(),
                "attribute": 0,
                "flag": 0,
            }
        )

    json_data["tracks"] = new_tracks
    json_data["duration"] = max(int(json_data.get("duration", 0)), cursor_us)

    input_path = Path("draft_content.json")
    backup_path = Path("draft_content.json.bak")
    if input_path.exists():
        backup_path.write_text(input_path.read_text(encoding="utf-8"), encoding="utf-8")
        logging.info(f"\nĐã tạo backup: {backup_path}")

    with open(input_path, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)

    logging.info(f"✅ Xử lý thành công! Đã đồng bộ và tái xây dựng timeline.")
    logging.info(
        "   (Chỉ làm chậm video khi audio dài hơn, giữ nguyên tốc độ khi audio ngắn hơn)"
    )
    logging.info(
        "   📝 Text: Thời gian bắt đầu = audio hiện tại, thời gian kết thúc = audio tiếp theo"
    )
    return True


def apply_style_only(json_file: str) -> bool:
    """
    CHỨC NĂNG MỚI: Chỉ áp dụng style cho TẤT CẢ text mà không thay đổi nội dung.
    CẬP NHẬT: Áp dụng cho toàn bộ text của tất cả các track text, không chỉ riêng vi-VN.
    """
    print(f"\n--- CHỨC NĂNG: CHỈ ÁP DỤNG STYLE ---")
    print(f"📁 Đang đọc file: {json_file}")
    print(f"🎨 Sẽ áp dụng style (font, size, màu sắc...) cho TẤT CẢ text")
    print(f"📝 KHÔNG thay đổi nội dung text, chỉ nâng cấp cấu trúc")

    # Đọc file JSON
    json_data = load_json_file(json_file)
    if not json_data:
        print(f"❌ Không thể đọc file '{json_file}'")
        return False

    # Kiểm tra có text nào không
    materials = json_data.get("materials", {})
    texts = materials.get("texts", [])

    if not texts:
        print("⚠️ Không tìm thấy text nào trong dự án.")
        return False

    print(f"📊 Tìm thấy {len(texts)} text để áp dụng style")

    # Áp dụng style
    if apply_style_to_texts(json_data):
        # Tạo backup trước khi lưu
        backup_path = json_file + ".bak"
        try:
            if os.path.exists(json_file):
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                os.rename(json_file, backup_path)
                print(f"💾 Đã tạo backup: {backup_path}")
        except Exception as e:
            print(f"⚠️ Cảnh báo: Không thể tạo backup: {e}")

        # Lưu file với style đã áp dụng
        try:
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(json_data, f, ensure_ascii=False, indent=4)
            print(f"✅ Đã lưu file với style vào: {json_file}")
            return True
        except Exception as e:
            print(f"❌ Lỗi khi lưu file {json_file}: {e}")
            return False
    else:
        print("❌ Lỗi trong quá trình áp dụng style.")
        return False


# ==============================================================================
# LOGIC FUNCTIONS FOR V7 METHODS
# ==============================================================================


def v7_export_captions_from_csv_logic(csv_file, json_file, output_dir):
    try:
        # Xuất caption vi-VN
        export_from_csv(csv_file, os.path.join(output_dir, "captions_vi.txt"))
        # Xuất caption tiếng Trung với giới hạn ký tự
        if json_file:
            results = get_translated_texts_with_timing(load_json_file(json_file))
        else:
            results = get_translated_texts_with_timing(
                load_json_file("draft_content.json")
            )
        export_chinese_with_char_count(
            results, os.path.join(output_dir, "captions_cn.txt")
        )
        return True
    except Exception as e:
        print(f"Lỗi: {e}")
        return False


def v7_replace_text_from_xlsx_logic(json_file, xlsx_file):
    return replace_vi_texts_in_draft_from_xlsx(json_file, xlsx_file)


def v7_apply_style_logic(json_file):
    return apply_style_only(json_file)


def v7_split_video_by_srt_logic(json_file, srt_file):
    json_data = load_json_file(json_file)
    if not json_data:
        return False
    return split_video_by_srt_timing(json_data, srt_file)


def v7_sync_video_audio_logic(json_file, video_start_index):
    json_data = load_json_file(json_file)
    if not json_data:
        return False
    return split_video_and_slow_down_for_audio(json_data, video_start_index)


def v7_split_by_character_logic(json_file, xlsx_file):
    json_data = load_json_file(json_file)
    if not json_data:
        return False
    return split_texts_into_multiple_tracks_by_character_index(json_data, xlsx_file)


def v7_export_cn_with_limit_logic(json_file, output_dir):
    try:
        results = get_translated_texts_with_timing(load_json_file(json_file))
        export_chinese_with_char_count(
            results, os.path.join(output_dir, "captions_cn_limit.txt")
        )
        return True
    except Exception as e:
        print(f"Lỗi: {e}")
        return False


def v7_export_to_srt(json_file: str, output_path: str) -> bool:
    """
    Xuất subtitle từ draft_content.json ra file SRT.
    Lấy text và timing từ extra_info > subtitle_fragment_info_list > sentence_list
    (Đây là nguồn dữ liệu chính xác mà CapCut sử dụng để xuất SRT)
    """
    try:
        # Đọc file JSON
        json_data = load_json_file(json_file)
        if not json_data:
            print(f"❌ Không thể đọc file '{json_file}'")
            return False
        
        # Lấy subtitle từ extra_info > subtitle_fragment_info_list
        srt_entries = []
        
        if 'extra_info' in json_data and json_data['extra_info'] is not None and isinstance(json_data['extra_info'], dict) and 'subtitle_fragment_info_list' in json_data['extra_info']:
            subtitle_fragments = json_data['extra_info']['subtitle_fragment_info_list']
            
            for fragment in subtitle_fragments:
                subtitle_cache_info = fragment.get('subtitle_cache_info', '')
                if not subtitle_cache_info:
                    continue
                
                try:
                    cache_info = json.loads(subtitle_cache_info)
                    if 'sentence_list' in cache_info and cache_info['sentence_list']:
                        for sentence in cache_info['sentence_list']:
                            # Lấy text từ sentence (text gốc, không phải translation)
                            text = sentence.get('text', '').strip()
                            if not text:
                                continue
                            
                            # Lấy timing từ sentence (đã là milliseconds)
                            start_time_ms = sentence.get('start_time', 0)
                            end_time_ms = sentence.get('end_time', 0)
                            
                            # Đảm bảo timing hợp lệ
                            if start_time_ms >= 0 and end_time_ms > start_time_ms:
                                # Loại bỏ HTML tags nếu có
                                clean_text = clean_text_from_html(text)
                                if clean_text:
                                    srt_entries.append({
                                        'start_time': int(start_time_ms),
                                        'end_time': int(end_time_ms),
                                        'text': clean_text
                                    })
                except Exception as e:
                    # Bỏ qua fragment lỗi, tiếp tục với fragment tiếp theo
                    continue
        
        if not srt_entries:
            print("❌ Không tìm thấy subtitle nào từ extra_info > subtitle_fragment_info_list")
            # Thử fallback: lấy từ text tracks nếu không có subtitle fragments
            print("   Đang thử lấy từ text tracks...")
            
            # Lấy text từ materials > texts
            materials = json_data.get('materials', {})
            texts_map = {}
            
            if materials and isinstance(materials, dict) and 'texts' in materials:
                for text_item in materials['texts']:
                    text_id = text_item.get('id', '')
                    if not text_id:
                        continue
                    
                    content = text_item.get('content', '')
                    text_content = extract_text_from_content(content)
                    
                    if not text_content:
                        recognize_text = text_item.get('recognize_text', '')
                        base_content = text_item.get('base_content', '')
                        text_content = clean_text_from_html(recognize_text) or clean_text_from_html(base_content)
                    
                    if text_content:
                        texts_map[text_id] = text_content.strip()
            
            # Lấy từ text tracks
            text_segments = []
            tracks = json_data.get('tracks', [])
            for track in tracks:
                if track.get('type') == 'text' and 'segments' in track:
                    text_segments.extend(track['segments'])
            
            text_segments.sort(key=lambda s: s.get('target_timerange', {}).get('start', 0))
            
            for segment in text_segments:
                material_id = segment.get('material_id', '')
                target_timerange = segment.get('target_timerange', {})
                start_us = target_timerange.get('start', 0)
                duration_us = target_timerange.get('duration', 0)
                end_us = start_us + duration_us
                
                text_content = texts_map.get(material_id, '')
                if not text_content:
                    segment_content = segment.get('content', '') or segment.get('text', '')
                    if segment_content:
                        text_content = extract_text_from_content(str(segment_content))
                
                if text_content:
                    start_ms = start_us // 1000
                    end_ms = end_us // 1000
                    if start_ms >= 0 and end_ms > start_ms:
                        srt_entries.append({
                            'start_time': start_ms,
                            'end_time': end_ms,
                            'text': text_content
                        })
        
        if not srt_entries:
            print("❌ Không tìm thấy subtitle nào để xuất")
            return False
        
        # Sắp xếp theo thời gian bắt đầu
        srt_entries.sort(key=lambda x: x['start_time'])
        
        # Ghi ra file SRT
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                for idx, entry in enumerate(srt_entries, 1):
                    # Số thứ tự
                    f.write(f"{idx}\n")
                    
                    # Thời gian: HH:MM:SS,mmm --> HH:MM:SS,mmm
                    start_time_str = milliseconds_to_srt_time(entry['start_time'])
                    end_time_str = milliseconds_to_srt_time(entry['end_time'])
                    f.write(f"{start_time_str} --> {end_time_str}\n")
                    
                    # Nội dung subtitle
                    f.write(f"{entry['text']}\n")
                    
                    # Dòng trống giữa các entry
                    f.write("\n")
            
            print(f"✅ Đã xuất {len(srt_entries)} subtitle vào file: {output_path}")
            return True
        except Exception as e:
            print(f"❌ Lỗi khi ghi file SRT: {e}")
            return False
            
    except Exception as e:
        print(f"❌ Lỗi khi xử lý: {e}")
        import traceback
        traceback.print_exc()
        return False


def v7_export_csv_from_draft_logic(json_file, output_dir):
    try:
        json_data = load_json_file(json_file)
        if not json_data:
            return False
        # Lấy thông tin text đã được translate
        results = get_translated_texts_with_timing(json_data)

        if not results:
            return False

        # Lưu kết quả ra file CSV
        csv_file = os.path.join(output_dir, "translated_texts.csv")
        save_to_csv(results, csv_file)

        # Kiểm tra và tạo file captions.xlsx nếu chưa có
        xlsx_file = os.path.join(output_dir, "captions.xlsx")
        create_captions_xlsx_if_not_exists(results, xlsx_file)

        return True
    except Exception as e:
        print(f"Lỗi: {e}")
        return False


def v7_find_idSubtile_and_nameAudio_sort(json_file, output_dir):
    """
    Lấy danh sách tên audio (là các text có audio đính kèm) từ draft_content.json.
    Sắp xếp theo thứ tự subtitle từ extra_info.
    """
    try:
        # Đọc file JSON
        json_data = load_json_file(json_file)
        if not json_data:
            print(f"❌ Không thể đọc file '{json_file}'")
            return False
        
        # Lấy danh sách audio materials trước
        materials = json_data.get('materials', {})
        audios = materials.get('audios', [])
        print(f"Debug: Found {len(audios)} audio materials")
        
        # Tạo mapping material_id -> text_id từ audios
        material_to_text_id = {}
        for audio in audios:
            if isinstance(audio, dict):
                material_id = audio.get('id', '')
                text_id = audio.get('text_id', '')
                if material_id and text_id:
                    material_to_text_id[material_id] = text_id
        
        # Lấy thứ tự subtitle từ extra_info
        subtitle_order = []
        if (
            "extra_info" in json_data
            and json_data["extra_info"]
            and isinstance(json_data["extra_info"], dict)
        ):
            if "subtitle_fragment_info_list" in json_data["extra_info"]:
                fragments = json_data["extra_info"]["subtitle_fragment_info_list"]
                print(f"Debug: Found {len(fragments)} subtitle fragments")
                for fragment in fragments:
                    if "subtitle_cache_info" in fragment and fragment["subtitle_cache_info"]:
                        try:
                            cache_info = json.loads(fragment["subtitle_cache_info"])
                            if "sentence_list" in cache_info:
                                for sentence in cache_info["sentence_list"]:
                                    text_id = sentence.get("text_id", "")
                                    if text_id:
                                        subtitle_order.append(text_id)
                        except:
                            continue
        
        print(f"Debug: Found {len(subtitle_order)} subtitle text_ids")
        
        # Nếu không có subtitle_order, fallback: lấy theo thứ tự segments trong audio tracks
        if not subtitle_order:
            print("Debug: No subtitle order found, using audio track segments order as fallback")
            tracks = json_data.get('tracks', [])
            audio_segments = []
            for track in tracks:
                if track.get('type') == 'audio':
                    segments = track.get('segments', [])
                    for seg in segments:
                        material_id = seg.get('material_id', '')
                        start_time = seg.get('target_timerange', {}).get('start', 0)
                        if material_id in material_to_text_id:
                            audio_segments.append((start_time, material_to_text_id[material_id]))
            
            # Sắp xếp theo start_time
            audio_segments.sort(key=lambda x: x[0])
            subtitle_order = [text_id for _, text_id in audio_segments]
            print(f"Debug: Using {len(subtitle_order)} audio text_ids from track segments as order")
        
        # Tạo mapping text_id -> audio name
        audio_mapping = {}
        for audio in audios:
            if isinstance(audio, dict):
                path = audio.get('path', '')
                text_id = audio.get('text_id', '')
                if '/textReading/' in path and '.wav' in path and text_id:
                    # Trích xuất tên đầy đủ từ path
                    start = path.find('/textReading/') + len('/textReading/')
                    end = path.find('.wav')
                    if start != -1 and end != -1:
                        full_name = path[start:end]
                        audio_mapping[text_id] = full_name
        
        logging.info(f"Debug: Found {len(audio_mapping)} audio mappings")
        
        # Sắp xếp theo thứ tự subtitle
        audio_names = []
        for text_id in subtitle_order:
            if text_id in audio_mapping:
                audio_names.append(audio_mapping[text_id])
        
        logging.info(f"Debug: Matched {len(audio_names)} audio names from {len(subtitle_order)} subtitles")
        
        if not audio_names:
            logging.warning("❌ Không tìm thấy audio nào khớp với subtitles")
            return False
        
        # Xuất ra file
        output_file = os.path.join(output_dir, 'audio_names_sorted.txt')
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                for name in audio_names:
                    f.write(name + '\n')
            logging.info(f"✅ Đã xuất {len(audio_names)} tên audio theo thứ tự subtitle vào file: {output_file}")
            return output_file  # Trả về đường dẫn file thay vì True
        except Exception as e:
            logging.error(f"❌ Lỗi khi ghi file: {e}")
            return False
            
    except Exception as e:
        logging.error(f"❌ Lỗi khi xử lý: {e}", exc_info=True)
        return False